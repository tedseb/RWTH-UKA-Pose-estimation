import numpy as np
import pickle
import cv2
import matplotlib.pyplot as plt
from pathlib import Path
import openpyxl


KNIEBEUGE_FRONT_SETS = {
    "data_path": "data/awinda/frame_diffs/kniebeuge_front.pickle",
    "rotate": False,
    "mappings": {
        0: (354, 1742),
        1: (356, 496),
        2: (604, 712),
        3: (712, 852),
        4: (852, 982),
        5: (982, 1098),
        6: (1098, 1220),
        7: (1120, 1354),
        8: (1354, 1480),
        9: (1480, 1614),
        10: (1614, 1740),
    },
}

KNIEBEUGE_SIDE_SETS = {
    "data_path": "data/awinda/frame_diffs/kniebeuge_side.pickle",
    "rotate": False,
    "mappings": {
        0: (520, 2076),
        1: (520, 636),
        2: (636, 768),
        3: (768, 922),
        4: (922, 1040),
        5: (1040, 1180),
        6: (1180, 1316),
        7: (1316, 1412),
        8: (1456, 1556),
        9: (1556, 1684),
        10: (1684, 1802),
        11: (1802, 1956),
        12: (1956, 2076),
    },
}

STEHEN_FRONT_SETS = {
    "data_path": "data/awinda/frame_diffs/stehen.pickle",
    "rotate": False,
    "mappings": {
        0: (648, 926),
    },
}

STEHEN_SIDE_SETS = {
    "data_path": "data/awinda/frame_diffs/stehen.pickle",
    "rotate": False,
    "mappings": {
        1: (1056, 1264),
        2: (1644, 1818),
        3: (1066, 1254),
    },
}

GEHEN_FRONT_SETS = {
    "data_path": "data/awinda/frame_diffs/gehen_front.pickle",
    "rotate": False,
    "mappings": {
        1: (326, 546),
        2: (846, 1060),
    },
}

GEHEN_LEFT_FRONT_SETS = {
    "data_path": "data/awinda/frame_diffs/gehen_left_front.pickle",
    "rotate": True,
    "mappings": {
        0: (86, 210),
    },
}

GEHEN_RIGHT_FRONT_SETS = {
    "data_path": "data/awinda/frame_diffs/gehen_right_front.pickle",
    "rotate": True,
    "mappings": {
        0: (82, 192),
    },
}

HOCKE_FRONT_SETS = {
    "data_path": "data/awinda/frame_diffs/kniebeuge_front.pickle",
    "mappings": {
        1: (440, 468),
        2: (644, 664),
        3: (1158, 1188),
        4: (1656, 1686),
    },
}

HOCKE_SIDE_SETS = {
    "data_path": "data/awinda/frame_diffs/kniebeuge_side.pickle",
    "mappings": {
        1: (686, 740),
        2: (944, 994),
        3: (1474, 1520),
        4: (1612, 1642),
        5: (1730, 1770),
    },
}

VIDEO_CONFIG_04_2023_KNIEBEUGEN = {
    "data_path": "04_2023_kniebeeugen.pickle",
    "mappings": {
        0: (170, 578),
        1: (198, 388),
        2: (388, 578),
        3: (1314, 1510),
        4: (1510, 1682),
        5: (1682, 1898),
    },
}

VIDEO_CONFIG_04_2023_LIEGESTUTZEN = {
    "data_path": "04_2023_liegestutzen.pickle",
    "mappings": {
        0: (168, 522),
        1: (168, 252),
        2: (252, 312),
        3: (452, 522),
    },
}

VIDEO_CONFIG_04_2023_SITUPS = {
    "data_path": "04_2023_situps.pickle",
    "mappings": {
        0: (148, 990),
        1: (148, 242),
        2: (408, 558),
        3: (948, 1256),
    },
}


VIDEO_CONFIG_05_2023_KNIEBEUGEN_FARBIG = {
    "data_path": "05_2023_kniebeeugen_farbig.pickle",
    "mappings": {
        0: (58, 300),
        1: (58, 182),
        2: (182, 300),
        3: (670, 790),
        4: (790, 926),
    },
}

VIDEO_CONFIG_05_2023_KNIEBEUGEN_SCHWARZ = {
    "data_path": "05_2023_kniebeeugen_schwarz.pickle",
    "mappings": {
        0: (130, 528),
        1: (130, 250),
        2: (250, 394),
        3: (674, 834),
        4: (834, 966),
    },
}

EVALUATION_LIST = {
    "kniebeugen_Shawan": VIDEO_CONFIG_04_2023_KNIEBEUGEN,
    "liegestutzen_Shawan": VIDEO_CONFIG_04_2023_LIEGESTUTZEN,
    "situps_Shawan": VIDEO_CONFIG_04_2023_SITUPS,
    "kniebeugen_schwarz_Alice": VIDEO_CONFIG_05_2023_KNIEBEUGEN_SCHWARZ,
    "kniebeugen_farbig_Alice": VIDEO_CONFIG_05_2023_KNIEBEUGEN_FARBIG,
}

# TODO: Evil hack. Don't do that!!!!
prefix = "plots"

class EvaluationCollector():
    def __init__(self):
        self._data = {}

    def add_data_frame(self, exercise_name: str, data: dict):
        self._data[exercise_name] = data

    def write_excel(self):
        wb = openpyxl.Workbook()
        sheet_obj = wb.active

        headers = [
            "Gelenk",
            "median abw. algo",
            "avg abw. algo",
            "max abw. algo",
            "median abw. Skelett",
            "avg abw. Skelett",
            "max abw. Skelett",
            "median abw. Winkel",
            "avg abw. Winkel",
            "max abw. Winkel",
        ]
        for i, header in enumerate(headers, 1):
            sheet_obj.cell(row=1, column=i).value = header
            letter = openpyxl.utils.get_column_letter(i)
            sheet_obj.column_dimensions[letter].width = 18

        row = 1
        for exercise, angle_data in self._data.items():
            print()
            row += 1

            print(exercise)
            sheet_obj.cell(row=row, column=1).value = exercise
            row += 1

            for angle, data in angle_data.items():
                print(angle, ": ", data)
                for i, name in enumerate([angle, *data], 1):
                    name = f"{name:.4f}" if isinstance(name, float) else name
                    sheet_obj.cell(row=row, column=i).value = name
                row += 1

        wb.save("./evaluation.xlsx")



def plot_graph(exercise, exercise_repetition, diff_angles, path_prefix="angle_diff", y_label="Flexion in Grad", use_title=True):
    y_values = [diff_angle[2] for diff_angle in diff_angles]
    x_values = [i / 30. for i in range(len(y_values))]
    plt.plot(x_values, y_values)
    plt.ylabel(y_label)
    plt.xlabel("Dauer in Sekunden")
    plt.subplots_adjust(bottom=0.15)
    if use_title:
        plt.title(f"{exercise}")  # {exercise_repetition}")
    title = f"./data/awinda/plot_{prefix}/{path_prefix}/{exercise}_{exercise_repetition}.png"
    title.replace(" ", "_")
    plt.savefig(title)
    plt.clf()


def compute_and_print_repetition_evaluation(exercise, exercise_repetition, repetition_interval, exercise_data, collector):
    angle_data = {}
    for angle_name in exercise_data["diff"]:
        # print(angle_name)
        diff = exercise_data["diff"][angle_name]
        # print(angle_name, repetition_interval)
        metrabs = exercise_data["metrabs"][angle_name]
        xsens = exercise_data["xsens"][angle_name]
        real_xsens = exercise_data["real_angle_xsens"][angle_name]

        diff_angles = []
        diff_skeletons = []
        metrabs_angles = []
        xsens_angles = []
        real_diff_angles = []
        for i in range(repetition_interval[0] + 1, repetition_interval[1] + 1, 2):
            if i in diff:
                metrabs_angle = metrabs[i]
                xsens_angle = xsens[i]
                diff_angles.append(np.abs(real_xsens[i] - metrabs_angle))
                diff_skeletons.append(np.abs(xsens_angle - metrabs_angle))
                metrabs_angles.append(metrabs_angle)
                xsens_angles.append(xsens_angle)
                real_diff_angles.append(np.absolute(xsens[i] - real_xsens[i]))

        median_real_angles = np.median(real_diff_angles, axis=0)
        max_real_angles = np.max(real_diff_angles, axis=0)
        avg_real_angles = np.average(real_diff_angles, axis=0)

        median_skeletons = np.median(diff_skeletons, axis=0)
        max_skeletons = np.max(diff_skeletons, axis=0)
        avg_skeletons = np.average(diff_skeletons, axis=0)

        median_angles = np.median(diff_angles, axis=0)
        max_angles = np.max(diff_angles, axis=0)
        avg_angles = np.average(diff_angles, axis=0)

        plt.rcParams.update({'font.size': 12})
        plot_graph(f"{exercise}-{angle_name}_difference", exercise_repetition, diff_skeletons, y_label="Differenz in Grad", use_title=True)
        plot_graph(f"{exercise}-{angle_name} RehaPlus System", exercise_repetition, metrabs_angles, "metrabs")
        plot_graph(f"{exercise}-{angle_name} Xsens System", exercise_repetition, xsens_angles, "xsens")
        plot_graph(f"{exercise}-{angle_name} Algorithm Error", exercise_repetition, real_diff_angles, "algorithm_error")



        # print("  Average:", f"X={average_angles[0]}", f"Y={average_angles[1]}", f"Z={average_angles[2]}")
        # print("  Median:", f"X={median_angles[0]}", f"Y={median_angles[1]}", f"Z={median_angles[2]}")
        # print("  Min:", f"X={min_angles[0]}", f"Y={min_angles[1]}", f"Z={min_angles[2]}")
        # print("  Max:", f"X={max_angles[0]}", f"Y={max_angles[1]}", f"Z={max_angles[2]}")
        angle_data[angle_name] = (
            median_real_angles[2],
            avg_real_angles[2],
            max_real_angles[2],
            median_skeletons[2],
            avg_skeletons[2],
            max_skeletons[2],
            median_angles[2],
            max_angles[2],
            avg_angles[2],
        )
        # # TODO:: Ugly Code
        # for i, diff_rots in diff.items():
        #     if diff_rots[2] == max_angles[2] and repetition_interval[0] <= i <= repetition_interval[1]:
        #         print("  Max Z Frame", i)
        #         return i

    # with open(f"./data/awinda/plot_{prefix}/summary_data/{exercise}_{exercise_repetition}.csv", 'w') as the_file:
    #     the_file.write("Gelenk, avg Abweichung der Berechnung, avg abweichung xsens-metrabs, median abweichung xsens-metrabs, min abweichung, max abweichung\n")
    #     print("Gelenk, avg Abweichung der Berechnung, avg abweichung xsens-metrabs, median abweichung xsens-metrabs, min abweichung, max abweichung")
    #     for name, data in angle_data.items():
    #         the_file.write(f"{name}, {data[4]}, {data[0]}, {data[1]}, {data[2]}, {data[3]}\n")
    #         print(f"{name}, {data[4]}, {data[5]}")

    if exercise_repetition == 0:
        print(f"  # Complete Set - ({repetition_interval}) #")
        exercise_repetition = "multiple"
    else:
        print(f"  # Repetition {exercise_repetition} - Interval ({repetition_interval}) #")

    collector.add_data_frame(f"{exercise}_{exercise_repetition}", angle_data)



def evaluate_exercise(exercise, exercise_repetitions, collector):
    print(f"### Evaluation Exercise - {exercise}")
    exercise_data = pickle.load(open(exercise_repetitions["data_path"], "rb"))
    # print(exercise_data)
    # cap = cv2.VideoCapture(exercise_data["path_video"])
    angle_data = {}
    for exercise_repetition, repetition_interval in exercise_repetitions["mappings"].items():
        max_z_frame = compute_and_print_repetition_evaluation(
            exercise, exercise_repetition, repetition_interval, exercise_data, collector)


        # video_frame = (repetition_interval[0] + repetition_interval[1]) // 4  # / 2 / 2 (middle of interval and half the fps)
        # print(exercise_repetition)
        # cap.set(cv2.CAP_PROP_POS_FRAMES, video_frame)
        # ret, cur_frame = cap.read()
        # if exercise_repetitions.get("rotate", False):
        #     cur_frame = cv2.rotate(cur_frame, cv2.ROTATE_90_COUNTERCLOCKWISE)
        # if ret:
        #     print("Screenshot")
        #     exercise_file_name = exercise.replace(" ", "_")
        #     exercise_file_name = f"{exercise_file_name}_{exercise_repetition}"
        #     cv2.imwrite(f"data/awinda/screenshots/evaluation/{exercise_file_name}.jpg", cur_frame)
        #     # print(f"Save data/awinda/screenshots/evaluation/{exercise_file_name}.jpg")
        # print()


def main():
    np.set_printoptions(suppress=True)
    np.set_printoptions(precision=3)
    collector = EvaluationCollector()
    global prefix

    for exercise, exercise_repetitions in EVALUATION_LIST.items():
        prefix = exercise
        for path_prefix in ["metrabs", "xsens", "algorithm_error", "angle_diff", "summary_data"]:
            Path(f"./data/awinda/plot_{prefix}/{path_prefix}").mkdir(parents=True, exist_ok=True)
        evaluate_exercise(exercise, exercise_repetitions, collector)
    collector.write_excel()


if __name__ == "__main__":
    main()
